from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .forms import StaffCreationForm,UserUpdateForm,\
TeacherSubjectAssignmentForm,StaffUploadForm
from django.http import HttpResponse, HttpResponseBadRequest
from schools.models import SchoolRegistration, Branch
from classes.models import TeacherSubjectClassAssignment,Subject
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage,PageNotAnInteger
from django.db.models import Q
from .models import Staff,Role
from utils.decorator import login_required_with_short_code
from django.db import transaction
import pandas as pd
from django.contrib.auth.models import User
import openpyxl
from django.core.mail import send_mail
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.worksheet.datavalidation import DataValidation
import zipfile, re
from io import BytesIO
from django.contrib.auth.models import User

def extract_branch_from_filename(filename, school):
    # Extract the school type and branch name from the filename using regex
    match = re.search(r'^(Primary|College)_(.*?)_staff_template\.xlsx$', filename, re.IGNORECASE)
    
    if match:
        primary_or_college = match.group(1)  # This captures "Primary" or "College"
        branch_name = normalize_branch_name_for_matching(match.group(2).strip().lower())
        
        print(f"Extracted branch name for matching: {branch_name}")  # Debugging

        # Attempt to find a branch by name (case-insensitive match)
        branch = Branch.objects.filter(branch_name__iexact=branch_name, school=school).first()
        
        if branch:
            print(f"Exact match found: {branch.branch_name}")  # Debugging
            return branch
        else:
            print(f"No exact match found for branch: {branch_name}")

    print("No match found")  # Debugging
    return None

def generate_unique_username(last_name, school_name):
    # Extract the first 3 letters of the school's name in uppercase
    school_prefix = school_name[:3].upper()
    # Ensure the last name is fully capitalized
    last_name = last_name.upper()
    base_username = f"{school_prefix}/{last_name}/"
    counter = 1

    # Find the next available unique username
    while True:
        username = f"{base_username}{counter}"  # Natural counter (1, 2, 3, etc.)
        if not User.objects.filter(username=username).exists():
            break
        counter += 1

    return username


@login_required_with_short_code
@transaction.atomic
def upload_staff(request, short_code):
    school = get_object_or_404(SchoolRegistration, short_code=short_code)

    if request.method == 'POST':
        form = StaffUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['file']
            try:
                if uploaded_file.name.endswith('.zip'):
                    with zipfile.ZipFile(uploaded_file, 'r') as zip_file:
                        for file_name in zip_file.namelist():
                            with zip_file.open(file_name) as excel_file:
                                branch = extract_branch_from_filename(file_name, school)  # Extract the branch
                                if branch:
                                    process_uploaded_file(excel_file, file_name, branch, school)
                else:
                    branch = extract_branch_from_filename(uploaded_file.name, school)  # Extract the branch
                    if branch:
                        process_uploaded_file(uploaded_file, uploaded_file.name, branch, school)

                messages.success(request, "Staff data has been successfully uploaded and updated.")
                return redirect('staff_list', short_code=school.short_code)

            except Exception as e:
                messages.error(request, f"An error occurred while processing the file: {str(e)}")
        else:
            messages.error(request, "Invalid form submission.")
    else:
        form = StaffUploadForm()

    return render(request, 'staff/upload_staff.html', {
        'form': form,
        'school': school,
    })

def process_uploaded_file(file, file_name, branch, school):
    # Load data from the second row of the sheet (after the header)
    if file_name.endswith('.xlsx'):
        data = pd.read_excel(file, header=2)  # Start reading from the third row (index 2)
    elif file_name.endswith('.csv'):
        data = pd.read_csv(file, header=2)  # Start reading from the third row (index 2)
    else:
        raise ValueError(f"Unsupported file type: {file_name}. Please upload an Excel or CSV file.")

    # Normalize the column names
    data.columns = data.columns.str.strip().str.lower()

    # Print normalized column names for debugging
    print(f"Normalized column names: {data.columns.tolist()}")

    # Ensure all columns are treated as strings where appropriate
    expected_columns = ['first_name', 'last_name', 'email', 'role', 'gender', 'marital_status', 
                        'date_of_birth', 'phone_number', 'address', 'nationality', 
                        'staff_category', 'status']

    for column in expected_columns:
        if column in data.columns:
            data[column] = data[column].astype(str).fillna('')

    # Print the first few rows to help with debugging
    print(f"Data read from file:\n{data.head()}")

    for idx, row in data.iterrows():
        try:
            first_name = row.get('first_name', '').strip()
            last_name = row.get('last_name', '').strip()

            # Log the index and data being processed
            print(f"Processing row {idx}: first_name='{first_name}', last_name='{last_name}'")

            # Skip rows with missing essential data
            if not first_name or not last_name:
                print(f"Skipping row {idx} due to missing first name or last name.")
                continue

            email = row.get('email', '').strip()
            role_name = row.get('role', '').strip()
            gender = row.get('gender', '').strip() if row.get('gender') else ''
            marital_status = row.get('marital_status', '').strip() if row.get('marital_status') else ''
            date_of_birth = row.get('date_of_birth', None)
            phone_number = str(row.get('phone_number', '')).strip() if row.get('phone_number') else ''
            address = row.get('address', '').strip() if row.get('address') else ''
            nationality = row.get('nationality', '').strip() if row.get('nationality') else ''
            staff_category = row.get('staff_category', '').strip() if row.get('staff_category') else ''
            status = row.get('status', '').strip() if row.get('status') else ''

            # Generate a unique username using the new format
            username = generate_unique_username(last_name, school.school_name)

            # Check if a user with this username exists
            if User.objects.filter(username=username).exists():
                print(f"User with username {username} already exists. Skipping.")
                continue

            # Create the user
            role, _ = Role.objects.get_or_create(name=role_name)
            user = User.objects.create_user(username=username, email=email, first_name=first_name, last_name=last_name.upper())
            user.set_password('new_staff')
            user.save()

            print(f"Created user: {username}")

            # Email the user with their login details (if email is provided)
            if email:
                send_mail(
                    subject='Your New Account',
                    message=f'Your account has been created. Your username is "{username}" and your password is "new_staff". Please log in and change your password.',
                    from_email='admin@example.com',
                    recipient_list=[email],
                )

            # Create or update the staff record
            staff, created = Staff.objects.update_or_create(user=user, defaults={
                'role': role,
                'gender': gender,
                'marital_status': marital_status,
                'date_of_birth': date_of_birth,
                'phone_number': phone_number,
                'address': address,
                'nationality': nationality,
                'staff_category': staff_category,
                'status': status,
            })

            # Assign the staff to the branch
            staff.branches.add(branch)
        
        except Exception as e:
            print(f"Error processing row {idx}: {e}")
            continue  # Skip to the next row if there's an error

def normalize_branch_name_for_matching(branch_name):
    # Convert to lowercase
    branch_name = branch_name.lower()
    # Replace spaces with underscores
    branch_name = branch_name.replace("_", " ")
    # Remove any special characters (like parentheses)
    branch_name = re.sub(r'[^\w\s]', '', branch_name)
    return branch_name

@login_required_with_short_code
def download_staff_template(request, short_code):
    school = get_object_or_404(SchoolRegistration, short_code=short_code)
    branches = Branch.objects.filter(school=school)

    if not branches.exists():
        return HttpResponseBadRequest("No branches found for this school.")

    files = []
    for branch in branches:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Staff Template"

        # Define the headers for the columns
        headers = [
            'first_name',
            'last_name',
            'email',
            'role',            # Dropdown for roles
            'gender',          # Dropdown for gender
            'marital_status',  # Dropdown for marital status
            'date_of_birth',
            'phone_number',
            'address',
            'nationality',     # Dropdown for nationality
            'staff_category',  # Dropdown for staff category
            'status'           # Dropdown for status
        ]

        # Style the school and branch header
        ws.merge_cells('A1:L1')
        ws['A1'] = f"{school.school_name} - {branch.branch_name}"
        ws['A1'].font = openpyxl.styles.Font(size=14, bold=True)
        ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')

        # Write the headers to the second row of the Excel sheet
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}2'] = header

        # Adjust column widths for readability
        column_widths = {
            'A': 15,  # first_name
            'B': 15,  # last_name
            'C': 30,  # email
            'D': 20,  # role
            'E': 10,  # gender
            'F': 15,  # marital_status
            'G': 15,  # date_of_birth
            'H': 20,  # phone_number
            'I': 30,  # address
            'J': 15,  # nationality
            'K': 20,  # staff_category
            'L': 10   # status
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Add demo rows to show users how to fill the data (sample data)
        demo_data_list = [
            ['John', 'Doe', 'john.doe@example.com', 'Teacher', 
            'male', 'single', '1990-01-01', '08012345678', 
            '123 Demo Street', 'nigerian', 'academic', 'active'],

            ['Jane', 'Smith', 'jane.smith@example.com', 'Administrator',
            'female', 'married', '1985-05-15', '09087654321',
            '456 Example Avenue', 'non_nigerian', 'non_academic', 'active'],

            ['Michael', 'Johnson', 'michael.johnson@example.com', 'Accountant',
            'male', 'divorced', '1982-09-20', '07023456789',
            '789 Sample Road', 'nigerian', 'academic', 'inactive'],
        ]

        # Add the demo rows to the worksheet
        for row_num, demo_data in enumerate(demo_data_list, start=3):
            for col_num, value in enumerate(demo_data, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = value

        # Fetch roles for dropdown
        roles = Role.objects.all().values_list('name', flat=True)

        # Define choices for dropdown fields
        gender_choices = [choice[0] for choice in Staff.GENDER_CHOICES]
        marital_status_choices = [choice[0] for choice in Staff.MARITAL_STATUS_CHOICES]
        nationality_choices = [choice[0] for choice in Staff.NATIONALITY_CHOICES]
        staff_category_choices = [choice[0] for choice in Staff.STAFF_CATEGORY_CHOICES]
        status_choices = [choice[0] for choice in Staff.STATUS_CHOICES]

        # Add roles and other choices to a new sheet for data validation dropdowns
        choices_sheet = wb.create_sheet(title="Choices")

        # Fill the roles
        for row_num, role in enumerate(roles, 1):
            choices_sheet[f'A{row_num}'] = role
        for row_num, choice in enumerate(gender_choices, 1):
            choices_sheet[f'B{row_num}'] = choice
        for row_num, choice in enumerate(marital_status_choices, 1):
            choices_sheet[f'C{row_num}'] = choice
        for row_num, choice in enumerate(nationality_choices, 1):
            choices_sheet[f'D{row_num}'] = choice
        for row_num, choice in enumerate(staff_category_choices, 1):
            choices_sheet[f'E{row_num}'] = choice
        for row_num, choice in enumerate(status_choices, 1):
            choices_sheet[f'F{row_num}'] = choice

        # Create a data validation dropdown for each field
        dv_roles = DataValidation(type="list", formula1="'Choices'!$A$1:$A$%d" % roles.count(), showDropDown=True)
        dv_gender = DataValidation(type="list", formula1="'Choices'!$B$1:$B$%d" % len(gender_choices), showDropDown=True)
        dv_marital_status = DataValidation(type="list", formula1="'Choices'!$C$1:$C$%d" % len(marital_status_choices), showDropDown=True)
        dv_nationality = DataValidation(type="list", formula1="'Choices'!$D$1:$D$%d" % len(nationality_choices), showDropDown=True)
        dv_staff_category = DataValidation(type="list", formula1="'Choices'!$E$1:$E$%d" % len(staff_category_choices), showDropDown=True)
        dv_status = DataValidation(type="list", formula1="'Choices'!$F$1:$F$%d" % len(status_choices), showDropDown=True)

        # Apply the data validation dropdown to the appropriate columns
        ws.add_data_validation(dv_roles)
        ws.add_data_validation(dv_gender)
        ws.add_data_validation(dv_marital_status)
        ws.add_data_validation(dv_nationality)
        ws.add_data_validation(dv_staff_category)
        ws.add_data_validation(dv_status)

        # Specify the columns for data validation
        dv_roles.add(f"D3:D1048576")
        dv_gender.add(f"E3:E1048576")
        dv_marital_status.add(f"F3:F1048576")
        dv_nationality.add(f"G3:G1048576")
        dv_staff_category.add(f"H3:H1048576")
        dv_status.add(f"I3:I1048576")

        # Normalize the branch name
        branch_name = normalize_branch_name_for_matching(branch.branch_name)

        # Determine the school type
        primary_or_college = "Primary" if branch.primary_school else "College"

        # Generate the filename using the normalized branch name and school type
        file_name = f"{primary_or_college}_{branch_name}_staff_template.xlsx"

        # Save the workbook to a BytesIO object
        file_buffer = BytesIO()
        wb.save(file_buffer)
        file_buffer.seek(0)
        
        # Store the file buffer and file name for zipping
        files.append((file_name, file_buffer))

    # If there's only one file, return it directly
    if len(files) == 1:
        response = HttpResponse(files[0][1], content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={files[0][0]}'
        return response

    # If multiple files, create a zip archive
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        for file_name, file_buffer in files:
            zip_file.writestr(file_name, file_buffer.getvalue())

    response = HttpResponse(zip_buffer.getvalue(), content_type="application/zip")
    response['Content-Disposition'] = f'attachment; filename={school.short_code}_staff_templates.zip'

    return response


@login_required_with_short_code
def add_staff(request, short_code):
    school = get_object_or_404(SchoolRegistration, short_code=short_code)
    branches = Branch.objects.filter(school=school)
    
    if request.method == 'POST':
        form = StaffCreationForm(request.POST, request.FILES, school=school)
        form.fields['branches'].queryset = branches
        if form.is_valid():
            user = form.save(commit=False)
            user.save()

            staff = Staff.objects.create(
                user=user,
                role=form.cleaned_data['role'],
                gender=form.cleaned_data['gender'],
                marital_status=form.cleaned_data['marital_status'],
                date_of_birth=form.cleaned_data['date_of_birth'],
                phone_number=form.cleaned_data['phone_number'],
                address=form.cleaned_data['address'],
                nationality=form.cleaned_data['nationality'],
                staff_category=form.cleaned_data['staff_category'],
                status=form.cleaned_data['status'],
                cv=form.cleaned_data.get('cv'),
                profile_picture=form.cleaned_data.get('profile_picture'),
                staff_signature=form.cleaned_data.get('staff_signature'),
            )

            staff.branches.set(form.cleaned_data['branches'])
            staff.save()

            messages.success(request, 'Staff member added successfully!')
            return redirect('staff_list', short_code=school.short_code)
        else:
            messages.error(request, 'Please correct the errors below and try again.')
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error in {field}: {error}")
    else:
        form = StaffCreationForm(school=school)
        form.fields['branches'].queryset = branches

    return render(request, 'staff/add_staff.html', {
        'form': form,
        'school': school,

    })


@login_required_with_short_code
def staff_list(request, short_code):
    school = get_object_or_404(SchoolRegistration, short_code=short_code)
    query = request.GET.get('q', '')
    per_page = request.GET.get('per_page', 10)
    status_filter = request.GET.get('status', '').lower()

    staff_members = Staff.objects.filter(branches__school=school).distinct()

    staff_members = staff_members.filter(
    Q(user__username__icontains=query) |
    Q(user__first_name__icontains=query) |
    Q(user__last_name__icontains=query) |
    Q(user__email__icontains=query) |
    Q(phone_number__icontains=query) |
    Q(role__name__icontains=query)
).distinct()

    if status_filter in ['active', 'inactive']:
        staff_members = staff_members.filter(status=status_filter)

    paginator = Paginator(staff_members, per_page)
    page = request.GET.get('page')
    try:
        staff_members = paginator.page(page)
    except PageNotAnInteger:
        staff_members = paginator.page(1)
    except EmptyPage:
        staff_members = paginator.page(paginator.num_pages)

    return render(request, 'staff/staff_list.html', {
        'staff_members': staff_members,
        'school': school,
        'query': query,
        'per_page': per_page,
        'status_filter': status_filter,
    })

@login_required_with_short_code
def edit_staff(request, short_code, staff_id):
    school = get_object_or_404(SchoolRegistration, short_code=short_code)
    staff = get_object_or_404(Staff, id=staff_id)

    branches = Branch.objects.filter(school=school)
    
    if request.method == 'POST':
        form = UserUpdateForm(request.POST, request.FILES, instance=staff.user, school=school)
        form.fields['branches'].queryset = branches
        if form.is_valid():
            # Save the form, user and staff instance will be updated
            form.save()

            messages.success(request, 'Staff member updated successfully!')
            return redirect('staff_list', short_code=school.short_code)
        else:
            messages.error(request, 'Please correct the errors below and try again.')
            # Add form errors to the messages framework
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error in {field}: {error}")
            
    else:
        form = UserUpdateForm(instance=staff.user, school=school)
        form.fields['branches'].queryset = branches
        form.fields['branches'].initial = staff.branches.all()

    return render(request, 'staff/edit_staff.html', {
        'form': form,
        'school': school,
        'staff': staff
    })

@login_required_with_short_code
def delete_staff(request, short_code, staff_id):
    school = get_object_or_404(SchoolRegistration, short_code=short_code)

    # Get the specific staff member based on ID
    staff = get_object_or_404(Staff, id=staff_id)

    # Ensure that the staff member is associated with the correct school branches
    if not staff.branches.filter(school=school).exists():
        messages.error(request, 'This staff member does not belong to this school.')
        return redirect('staff_list', short_code=short_code)

    if request.method == 'POST':
        user = staff.user  # Capture the associated user before deleting staff
        staff.delete()
        # Optionally, delete the user if required
        user.delete()

        messages.success(request, 'Staff member deleted successfully.')
        return redirect('staff_list', short_code=short_code)
    
    return render(request, 'staff/delete_staff.html', {
        'school': school,
        'staff': staff
    })


@transaction.atomic
@login_required_with_short_code
def assign_subjects_to_staff(request, short_code, staff_id):
    school = get_object_or_404(SchoolRegistration, short_code=short_code)
    staff = get_object_or_404(Staff, id=staff_id)
    branches = Branch.objects.filter(school=school, staff__id=staff.id).distinct()

    branch_id = request.POST.get('branch') or request.GET.get('branch')
    selected_branch = Branch.objects.get(id=branch_id) if branch_id else None

    if request.method == 'POST':
        form = TeacherSubjectAssignmentForm(request.POST, school=school, branch_id=branch_id, staff=staff)
        if form.is_valid():
            branch = form.cleaned_data['branch']
            subject = form.cleaned_data['subject']
            classes = form.cleaned_data['classes']

            # Unassign the subject from the selected classes for any other teacher in the same branch
            existing_assignments = TeacherSubjectClassAssignment.objects.filter(
                subject=subject,
                branch=branch,
                classes_assigned__in=classes
            ).exclude(teacher=staff)

            for assignment in existing_assignments:
                assignment.classes_assigned.remove(*classes)
                if assignment.classes_assigned.count() == 0:
                    assignment.delete()

            # Create or update the assignment for the current teacher in the specific branch
            assignment, created = TeacherSubjectClassAssignment.objects.get_or_create(
                teacher=staff,
                subject=subject,
                branch=branch
            )
            assignment.classes_assigned.add(*classes)
            assignment.save()

            messages.success(request, f'Subjects and classes assigned to {staff.user.first_name} successfully!')
            return redirect('teacher_assignments', short_code=school.short_code)
        else:
            print(f"Form Errors: {form.errors}")
    else:
        form = TeacherSubjectAssignmentForm(school=school, branch_id=branch_id, staff=staff)

    return render(request, 'staff/assign_subjects.html', {
        'form': form,
        'school': school,
        'staff': staff,
        'branches': branches,
        'selected_branch': selected_branch,
    })


@login_required_with_short_code
def teacher_assignments_view(request, short_code):
    school = get_object_or_404(SchoolRegistration, short_code=short_code)
    
    branches = Branch.objects.filter(school=school)
    
    selected_branch_id = request.GET.get('branch_id')
    search_query = request.GET.get('search', '')

    selected_branch = None
    teacher_assignments = {}
    page_obj = None  # Initialize page_obj as None

    if selected_branch_id:
        selected_branch = get_object_or_404(Branch, id=selected_branch_id, school=school)
        assignments = TeacherSubjectClassAssignment.objects.filter(branch=selected_branch).select_related('teacher', 'subject').prefetch_related('classes_assigned')
        
        if search_query:
            assignments = assignments.filter(teacher__user__first_name__icontains=search_query) | assignments.filter(teacher__user__last_name__icontains=search_query)
        
        # Pagination
        paginator = Paginator(assignments, 10)  # Show 10 assignments per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # Organize assignments by teacher
        for assignment in page_obj:
            teacher = assignment.teacher
            if teacher not in teacher_assignments:
                teacher_assignments[teacher] = []
            teacher_assignments[teacher].append({
                'subject': assignment.subject,
                'classes': assignment.classes_assigned.all()
            })

    return render(request, 'staff/teacher_assignments.html', {
        'school': school,
        'branches': branches,
        'selected_branch': selected_branch,
        'teacher_assignments': teacher_assignments,
        'page_obj': page_obj,
        'search_query': search_query
    })